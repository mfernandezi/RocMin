"""
Dashboard Unificado PyT - Los Pelambres
Streamlit App para análisis de perforación
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from datetime import time as dtime
import io

# Configuración de página
st.set_page_config(
    page_title="Dashboard PyT - Los Pelambres",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilos CSS personalizados
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #00B5B8, #EA6B2B);
        padding: 20px;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 20px;
    }
    .metric-card {
        background: white;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        border-left: 4px solid #00B5B8;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #f0f2f6;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #00B5B8 !important;
        color: white !important;
    }
    .positive { color: #22543d; font-weight: bold; }
    .negative { color: #c53030; font-weight: bold; }
</style>
""", unsafe_allow_html=True)

# Funciones auxiliares
def normalizar_equipo(rig):
    """Normaliza el nombre del equipo"""
    if not rig or pd.isna(rig):
        return ''
    n = str(rig).strip().upper().replace('-', '')
    if n == 'PFAR':
        n = 'PFARR'
    if n.startswith('PF') and len(n) == 3:
        n = 'PF0' + n[2]
    return n

def excel_date_to_datetime(val):
    """Convierte fecha de Excel a datetime"""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)) and 40000 < val < 50000:
        return datetime(1899, 12, 30) + timedelta(days=int(val))
    if isinstance(val, str):
        try:
            return pd.to_datetime(val)
        except:
            return None
    return None

def format_number(n, decimals=1):
    """Formatea número con separador de miles"""
    if n is None or pd.isna(n):
        return '-'
    return f"{n:,.{decimals}f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def agregar_dia_operacional_y_turno(df: pd.DataFrame, ts_col: str, out_date_col: str = "dia_operacional") -> pd.DataFrame:
    """
    Agrega columnas:
    - out_date_col: día operacional (si hora>=21 => día+1)
    - turno: 'A' si hora>=21 o hora<9, si no 'B'

    Supuesto operacional (según nombres de archivos típicos 21:00-09:00):
    - Turno A: 21:00-09:00
    - Turno B: 09:00-21:00
    """
    if ts_col not in df.columns:
        return df
    if len(df) == 0:
        df[out_date_col] = pd.Series(dtype="object")
        df["turno"] = pd.Series(dtype="object")
        return df

    ts = pd.to_datetime(df[ts_col], errors="coerce")
    hour = ts.dt.hour

    df = df.copy()
    df["turno"] = np.where((hour >= 21) | (hour < 9), "A", "B")
    base_day = ts.dt.floor("D")
    df[out_date_col] = (base_day + pd.to_timedelta((hour >= 21).astype(int), unit="D")).dt.date
    return df

@st.cache_data
def procesar_uebd(df, anio=None):
    """Procesa archivo UEBD - Versión optimizada con vectorización"""
    # Identificar columnas
    cols = df.columns.tolist()
    col_rig = 0
    col_fecha = 7 if len(cols) > 7 else 4
    col_duracion = 14 if len(cols) > 14 else len(cols) - 1
    col_codigo = 15 if len(cols) > 15 else len(cols) - 1
    col_estado = 16 if len(cols) > 16 else len(cols) - 1
    col_planificado = 19 if len(cols) > 19 else len(cols) - 1

    # Crear DataFrame con columnas renombradas
    result = pd.DataFrame()
    result['rig_raw'] = df.iloc[:, col_rig].astype(str).str.strip().str.upper().str.replace('-', '', regex=False)
    result['fecha_raw'] = pd.to_datetime(df.iloc[:, col_fecha], errors='coerce')
    result['duracion'] = pd.to_numeric(df.iloc[:, col_duracion], errors='coerce').fillna(0)
    result['codigo'] = df.iloc[:, col_codigo].astype(str).fillna('')
    result['estado'] = df.iloc[:, col_estado].astype(str).fillna('')
    result['planificado'] = df.iloc[:, col_planificado].astype(str).fillna('')

    # Normalizar equipos vectorizado
    result['rig'] = result['rig_raw'].apply(lambda x: 'PF0' + x[2] if x.startswith('PF') and len(x) == 3 else x)
    result.loc[result['rig'] == 'PFAR', 'rig'] = 'PFARR'

    # Filtrar equipos válidos y excluir PF03
    mask = (result['rig'].str.startswith('PF')) & (result['rig'] != 'PF03') & (result['fecha_raw'].notna())

    # Filtrar por año si se especifica
    if anio:
        mask = mask & (result['fecha_raw'].dt.year == anio)

    result = result[mask].copy()

    # Extraer componentes de fecha
    result['fecha'] = result['fecha_raw'].dt.date
    result['mes'] = result['fecha_raw'].dt.month
    result['anio'] = result['fecha_raw'].dt.year

    # Seleccionar columnas finales
    # Mantener timestamp para poder separar Turno A/B en reportes
    result['timestamp'] = result['fecha_raw']
    return result[['rig', 'fecha', 'mes', 'anio', 'timestamp', 'duracion', 'codigo', 'estado', 'planificado']].reset_index(drop=True)

@st.cache_data
def procesar_qaqc(df, anio=None):
    """Procesa archivo QAQC con desviaciones - Versión optimizada con vectorización"""
    cols = df.columns.tolist()

    # Identificar columnas - buscar por nombre
    col_map = {
        'rig': 0,
        'hole': 2,
        'metros': 26,
        'fecha': 29,
        'malla': None,
        'desv_xy': None,
        'desv_largo': None
    }

    # Buscar columnas por nombre
    for i, col in enumerate(cols):
        col_lower = str(col).lower()
        if 'rigname' in col_lower:
            col_map['rig'] = i
        elif 'holename' in col_lower:
            col_map['hole'] = i
        elif 'reallength' in col_lower:
            col_map['metros'] = i
        elif 'workdaystarted' in col_lower:
            col_map['fecha'] = i
        elif 'malla' in col_lower or 'mesh' in col_lower or 'pattern' in col_lower:
            col_map['malla'] = i
        elif 'desv' in col_lower and ('xy' in col_lower or 'horiz' in col_lower):
            col_map['desv_xy'] = i
        elif 'desv' in col_lower and ('largo' in col_lower or 'length' in col_lower or 'prof' in col_lower):
            col_map['desv_largo'] = i
        elif 'deviation' in col_lower and 'xy' in col_lower:
            col_map['desv_xy'] = i
        elif 'deviation' in col_lower and 'length' in col_lower:
            col_map['desv_largo'] = i

    # Crear DataFrame con columnas necesarias (vectorizado)
    result = pd.DataFrame()
    result['rig_raw'] = df.iloc[:, col_map['rig']].astype(str).str.strip().str.upper().str.replace('-', '', regex=False)
    result['fecha_raw'] = pd.to_datetime(df.iloc[:, col_map['fecha']], errors='coerce')
    result['metros'] = pd.to_numeric(df.iloc[:, col_map['metros']], errors='coerce').fillna(0)
    result['hole'] = df.iloc[:, col_map['hole']].astype(str).fillna('')

    # Columnas opcionales
    if col_map['malla'] is not None:
        result['malla'] = df.iloc[:, col_map['malla']].astype(str).fillna('Sin Malla')
    else:
        result['malla'] = 'Sin Malla'

    if col_map['desv_xy'] is not None:
        result['desv_xy'] = pd.to_numeric(df.iloc[:, col_map['desv_xy']], errors='coerce')
    else:
        result['desv_xy'] = np.nan

    if col_map['desv_largo'] is not None:
        result['desv_largo'] = pd.to_numeric(df.iloc[:, col_map['desv_largo']], errors='coerce')
    else:
        result['desv_largo'] = np.nan

    # Normalizar equipos vectorizado
    result['rig'] = result['rig_raw'].apply(lambda x: 'PF0' + x[2] if x.startswith('PF') and len(x) == 3 else x)
    result.loc[result['rig'] == 'PFAR', 'rig'] = 'PFARR'

    # Filtrar: equipos válidos, excluir PF03, metros válidos (0-100)
    mask = (
        (result['rig'].str.startswith('PF')) &
        (result['rig'] != 'PF03') &
        (result['fecha_raw'].notna()) &
        (result['metros'] > 0) &
        (result['metros'] <= 100)
    )

    # Filtrar por año si se especifica
    if anio:
        mask = mask & (result['fecha_raw'].dt.year == anio)

    result = result[mask].copy()

    # Extraer componentes de fecha
    result['fecha'] = result['fecha_raw'].dt.date
    result['mes'] = result['fecha_raw'].dt.month
    result['anio'] = result['fecha_raw'].dt.year

    # Seleccionar columnas finales
    # Mantener timestamp para poder separar Turno A/B en reportes
    result['timestamp'] = result['fecha_raw']
    return result[['rig', 'fecha', 'mes', 'anio', 'timestamp', 'hole', 'metros', 'malla', 'desv_xy', 'desv_largo']].reset_index(drop=True)

@st.cache_data
def procesar_plan_semanal(df):
    """
    Procesa archivo de plan semanal - Versión optimizada
    Estructura: Fila 0 = encabezados (Equipo, Índices, fechas...)
    """
    datos = {}
    cols = df.columns.tolist()

    # Función para convertir cualquier formato de fecha
    def parse_fecha_flexible(val):
        if pd.isna(val):
            return None
        # Número Excel (serial date)
        if isinstance(val, (int, float)) and 40000 < val < 50000:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
        # Ya es datetime
        if isinstance(val, datetime):
            return val.date()
        # String
        if isinstance(val, str):
            val = val.strip()
            # Formato ISO: 2025-01-15
            if len(val) >= 10 and val[4] == '-':
                try:
                    return datetime.strptime(val[:10], '%Y-%m-%d').date()
                except:
                    pass
            # Formato dd/mm/yyyy
            if '/' in val:
                try:
                    parts = val.split('/')
                    if len(parts) == 3:
                        if len(parts[2]) == 4:  # dd/mm/yyyy
                            return datetime(int(parts[2]), int(parts[1]), int(parts[0])).date()
                        elif len(parts[0]) == 4:  # yyyy/mm/dd
                            return datetime(int(parts[0]), int(parts[1]), int(parts[2])).date()
                except:
                    pass
            # Intentar pandas como último recurso
            try:
                return pd.to_datetime(val).date()
            except:
                pass
        return None

    # Identificar columnas de fecha (desde columna 2 en adelante)
    fechas_cols = {}
    for i, col in enumerate(cols[2:], start=2):
        fecha = parse_fecha_flexible(col)
        if fecha and fecha.year >= 2020:
            fechas_cols[i] = fecha

    if not fechas_cols:
        return datos

    # Convertir DataFrame a arrays para procesamiento más rápido
    equipos_raw = df.iloc[:, 0].astype(str).str.strip().str.upper().str.replace('-', '', regex=False)
    indices_raw = df.iloc[:, 1].astype(str).str.strip().str.lower()

    # Normalizar equipos
    equipos = equipos_raw.apply(lambda x: 'PF0' + x[2] if x.startswith('PF') and len(x) == 3 else x)
    equipos = equipos.replace('PFAR', 'PFARR')

    # Máscara de filas válidas
    mask_valido = equipos.str.startswith('PF') & (equipos != 'PF03') & (indices_raw != '') & (indices_raw != 'nan')

    # Procesar solo filas válidas
    for idx in df.index[mask_valido]:
        equipo = equipos[idx]
        indice = indices_raw[idx]

        for col_idx, fecha in fechas_cols.items():
            if fecha not in datos:
                datos[fecha] = {'total': {}, 'por_rig': {}}

            if equipo not in datos[fecha]['por_rig']:
                datos[fecha]['por_rig'][equipo] = {
                    'disponibilidad': 0, 'utilizacion': 0,
                    'rendimiento': 0, 'metros': 0, 'horas_efectivas': 0
                }

            try:
                val = float(df.iloc[idx, col_idx]) if pd.notna(df.iloc[idx, col_idx]) else 0
            except:
                val = 0

            d = datos[fecha]['por_rig'][equipo]

            if 'disponibilidad' in indice:
                d['disponibilidad'] = val * 100 if 0 < val < 2 else val
            elif 'utilizaci' in indice:
                d['utilizacion'] = val * 100 if 0 < val < 2 else val
            elif 'rendimiento' in indice:
                d['rendimiento'] = val
            elif indice == 'metros' or ('metro' in indice and 'hr' not in indice):
                d['metros'] = val
            elif 'horas' in indice and 'efect' in indice:
                d['horas_efectivas'] = val

    # Calcular totales por fecha (vectorizado)
    for fecha in datos:
        equipos_dict = datos[fecha]['por_rig']
        if not equipos_dict:
            continue

        vals = list(equipos_dict.values())
        sD = sum(v['disponibilidad'] for v in vals if v['disponibilidad'] > 0)
        cD = sum(1 for v in vals if v['disponibilidad'] > 0)
        sU = sum(v['utilizacion'] for v in vals if v['utilizacion'] > 0)
        cU = sum(1 for v in vals if v['utilizacion'] > 0)
        sR = sum(v['rendimiento'] for v in vals if v['rendimiento'] > 0)
        cR = sum(1 for v in vals if v['rendimiento'] > 0)
        sM = sum(v['metros'] for v in vals)
        sH = sum(v['horas_efectivas'] for v in vals)

        datos[fecha]['total'] = {
            'disponibilidad': sD / cD if cD > 0 else 0,
            'utilizacion': sU / cU if cU > 0 else 0,
            'rendimiento': sR / cR if cR > 0 else 0,
            'metros': sM,
            'horas_efectivas': sH
        }

    return datos

@st.cache_data
def procesar_plan_mensual(df):
    """Procesa archivo de plan mensual"""
    datos = {}
    cols = df.columns.tolist()

    # Buscar columnas de meses
    meses_nombres = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                     'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    meses_cols = {}

    for i, col in enumerate(cols):
        col_lower = str(col).lower()
        for j, mes in enumerate(meses_nombres, 1):
            if mes in col_lower or col_lower == mes[:3]:
                meses_cols[j] = i
                break

    # Si no encontró por nombre, asumir que están desde la columna 3
    if not meses_cols:
        for i in range(12):
            if 3 + i < len(cols):
                meses_cols[i + 1] = 3 + i

    equipo_actual = None
    for idx, row in df.iterrows():
        try:
            # Verificar si es una fila de equipo
            col0 = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ''
            if col0.startswith('PF') or col0 == 'PFAR':
                equipo_actual = normalizar_equipo(col0)

            if not equipo_actual or equipo_actual == 'PF03':
                continue

            indice = str(row.iloc[1]).strip().lower() if pd.notna(row.iloc[1]) else ''
            if not indice:
                continue

            if equipo_actual not in datos:
                datos[equipo_actual] = {m: {
                    'disponibilidad': 0, 'utilizacion': 0,
                    'rendimiento': 0, 'metros': 0, 'horas_efectivas': 0
                } for m in range(1, 13)}

            for mes, col_idx in meses_cols.items():
                val = float(row.iloc[col_idx]) if pd.notna(row.iloc[col_idx]) else 0
                d = datos[equipo_actual][mes]

                if 'disponibilidad' in indice:
                    d['disponibilidad'] = val * 100 if 0 < val < 2 else val
                elif 'utilizaci' in indice:
                    d['utilizacion'] = val * 100 if 0 < val < 2 else val
                elif 'rendimiento' in indice:
                    d['rendimiento'] = val
                elif indice == 'metros' or ('metro' in indice and 'hr' not in indice):
                    d['metros'] = val
                elif 'horas' in indice and 'efect' in indice:
                    d['horas_efectivas'] = val
        except:
            continue

    return datos

def calcular_metricas_uebd(df_uebd, df_qaqc, filtro_equipo=None, filtro_fecha_ini=None, filtro_fecha_fin=None):
    """Calcula métricas a partir de UEBD y QAQC - Versión optimizada vectorizada"""
    df_u = df_uebd
    df_q = df_qaqc

    # Aplicar filtros usando máscaras booleanas (más eficiente)
    if filtro_equipo and filtro_equipo != 'TODOS':
        df_u = df_u[df_u['rig'] == filtro_equipo]
        df_q = df_q[df_q['rig'] == filtro_equipo]

    if filtro_fecha_ini:
        df_u = df_u[df_u['fecha'] >= filtro_fecha_ini]
        df_q = df_q[df_q['fecha'] >= filtro_fecha_ini]

    if filtro_fecha_fin:
        df_u = df_u[df_u['fecha'] <= filtro_fecha_fin]
        df_q = df_q[df_q['fecha'] <= filtro_fecha_fin]

    # Calcular horas vectorizado (sin iterrows)
    if len(df_u) > 0:
        df_u = df_u.copy()
        df_u['horas'] = df_u['duracion'] / 3600

        # Máscaras para cada categoría
        mask_efectivo = df_u['estado'] == 'Efectivo'
        mask_perforando = mask_efectivo & df_u['codigo'].str.contains('106_Efectivo_Perforando', na=False)
        mask_mantencion = df_u['estado'] == 'Mantencion'
        mask_reserva = df_u['estado'] == 'Reserva'
        mask_demora = df_u['estado'] == 'Demora'
        mask_programada = df_u['planificado'] == 'Programada'

        horas = {
            'ef_total': df_u.loc[mask_efectivo, 'horas'].sum(),
            'ef_perf': df_u.loc[mask_perforando, 'horas'].sum(),
            'mant_prog': df_u.loc[mask_mantencion & mask_programada, 'horas'].sum(),
            'mant_no_prog': df_u.loc[mask_mantencion & ~mask_programada, 'horas'].sum(),
            'reserva': df_u.loc[mask_reserva, 'horas'].sum(),
            'dem_prog': df_u.loc[mask_demora & mask_programada, 'horas'].sum(),
            'dem_no_prog': df_u.loc[mask_demora & ~mask_programada, 'horas'].sum()
        }
    else:
        horas = {
            'ef_total': 0, 'ef_perf': 0, 'dem_prog': 0, 'dem_no_prog': 0,
            'reserva': 0, 'mant_prog': 0, 'mant_no_prog': 0
        }

    # Calcular metros
    metros = df_q['metros'].sum() if len(df_q) > 0 else 0
    pozos = df_q['hole'].nunique() if 'hole' in df_q.columns and len(df_q) > 0 else 0

    # Calcular indicadores
    horas_operativas = horas['ef_total'] + horas['reserva'] + horas['dem_prog'] + horas['dem_no_prog']
    horas_totales = horas_operativas + horas['mant_prog'] + horas['mant_no_prog']

    disponibilidad = (horas_operativas / horas_totales * 100) if horas_totales > 0 else 0
    uebd = (horas['ef_total'] / horas_operativas * 100) if horas_operativas > 0 else 0
    rend_efectivo = metros / horas['ef_perf'] if horas['ef_perf'] > 0 else 0
    rend_asarco = metros / horas['ef_total'] if horas['ef_total'] > 0 else 0

    return {
        'disponibilidad': disponibilidad,
        'uebd': uebd,
        'rend_efectivo': rend_efectivo,
        'rend_asarco': rend_asarco,
        'metros': metros,
        'pozos': pozos,
        'horas': horas,
        'horas_operativas': horas_operativas,
        'horas_totales': horas_totales
    }

@st.cache_data
def leer_archivo_optimizado(uploaded_file, tipo_archivo='auto', anio_filtro=None):
    """
    Lee archivo Excel o CSV de forma optimizada para archivos grandes.

    Optimizaciones:
    - CSV: Engine C, lectura por chunks si es muy grande, dtypes especificados
    - Excel: read_only mode con openpyxl para menor uso de memoria
    - Filtrado opcional por año durante la lectura
    """
    try:
        filename = uploaded_file.name.lower()
        file_size = uploaded_file.seek(0, 2)  # Obtener tamaño
        uploaded_file.seek(0)

        # Para archivos muy grandes (>50MB), usar chunks
        CHUNK_SIZE = 50000
        LARGE_FILE_THRESHOLD = 50 * 1024 * 1024  # 50MB

        if filename.endswith('.csv'):
            # CSV: detectar separador
            first_line = uploaded_file.readline().decode('utf-8', errors='ignore')
            uploaded_file.seek(0)
            sep = ';' if ';' in first_line else (',' if ',' in first_line else '\t')

            # Definir dtypes para evitar inferencia (más rápido)
            dtype_spec = {
                0: str,   # rig
                2: str,   # hole
                7: str,   # fecha
                14: float,  # duracion
                15: str,    # codigo
                16: str,    # estado
                19: str,    # planificado
                26: float,  # metros
                29: str,    # fecha qaqc
            }

            if file_size > LARGE_FILE_THRESHOLD:
                # Lectura por chunks para archivos muy grandes
                chunks = []
                for chunk in pd.read_csv(
                    uploaded_file,
                    sep=sep,
                    chunksize=CHUNK_SIZE,
                    encoding='utf-8',
                    on_bad_lines='skip',
                    low_memory=True
                ):
                    # Filtrar por año durante la lectura si se especifica
                    if anio_filtro:
                        # Intentar filtrar en columnas de fecha comunes
                        for col_idx in [7, 29, 37]:
                            if col_idx < len(chunk.columns):
                                try:
                                    chunk_dates = pd.to_datetime(chunk.iloc[:, col_idx], errors='coerce')
                                    mask = chunk_dates.dt.year == anio_filtro
                                    chunk = chunk[mask | chunk_dates.isna()]
                                    break
                                except:
                                    pass
                    chunks.append(chunk)
                return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
            else:
                # Lectura directa para archivos medianos
                return pd.read_csv(
                    uploaded_file,
                    sep=sep,
                    encoding='utf-8',
                    on_bad_lines='skip',
                    low_memory=True
                )
        else:
            # Excel: usar openpyxl con read_only para archivos grandes
            from openpyxl import load_workbook

            if file_size > LARGE_FILE_THRESHOLD:
                # Para archivos Excel grandes, usar read_only mode
                uploaded_file.seek(0)
                wb = load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active

                # Leer datos a lista para crear DataFrame
                data = []
                headers = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = row
                    else:
                        # Filtrar por año si se especifica
                        if anio_filtro:
                            # Buscar fecha en columnas comunes
                            fecha_val = row[7] if len(row) > 7 else (row[29] if len(row) > 29 else None)
                            if fecha_val:
                                try:
                                    fecha_dt = pd.to_datetime(fecha_val, errors='coerce')
                                    if pd.notna(fecha_dt) and fecha_dt.year != anio_filtro:
                                        continue
                                except:
                                    pass
                        data.append(row)

                wb.close()
                return pd.DataFrame(data, columns=headers) if data else pd.DataFrame()
            else:
                # Para archivos pequeños, lectura normal
                return pd.read_excel(
                    uploaded_file,
                    header=0,
                    engine='openpyxl'
                )
    except Exception as e:
        st.error(f"Error leyendo archivo: {e}")
        return None


# Inicializar session state
if 'df_uebd' not in st.session_state:
    st.session_state.df_uebd = None
if 'df_qaqc' not in st.session_state:
    st.session_state.df_qaqc = None
if 'plan_semanal' not in st.session_state:
    st.session_state.plan_semanal = None
if 'plan_mensual' not in st.session_state:
    st.session_state.plan_mensual = None

# Header
st.markdown("""
<div class="main-header">
    <h1>📊 Dashboard PyT Unificado</h1>
    <p>Los Pelambres - Análisis de Perforación</p>
</div>
""", unsafe_allow_html=True)

# Sidebar - Carga de archivos
with st.sidebar:
    st.header("📁 Carga de Archivos")

    anio_seleccionado = st.selectbox("Año de análisis",
                                      options=list(range(2026, 2019, -1)),
                                      index=1)

    st.subheader("Archivos de Datos")

    # Nota sobre formato CSV para archivos grandes
    st.caption("⚡ **Tip:** Para archivos grandes, usar CSV es más rápido. Si tu Excel tiene conexión externa, guárdalo como CSV primero.")

    file_uebd = st.file_uploader("📋 Archivo UEBD", type=['xlsx', 'xls', 'csv'], key='uebd')
    if file_uebd:
        with st.status("🔩 Cargando datos de perforación...", expanded=True) as status:
            st.write("⛏️ Extrayendo registros de tiempos...")
            # Pasar año para filtrar durante la lectura (más eficiente)
            df = leer_archivo_optimizado(file_uebd, anio_filtro=anio_seleccionado)
            if df is not None:
                st.write("🎯 Procesando datos por equipo...")
                st.session_state.df_uebd = procesar_uebd(df, anio_seleccionado)
                status.update(label=f"✅ UEBD cargado: {len(st.session_state.df_uebd):,} registros", state="complete")

    file_qaqc = st.file_uploader("📊 Archivo QAQC", type=['xlsx', 'xls', 'csv'], key='qaqc')
    if file_qaqc:
        with st.status("💎 Cargando datos de pozos...", expanded=True) as status:
            st.write("🕳️ Leyendo metros perforados...")
            # Pasar año para filtrar durante la lectura (más eficiente)
            df = leer_archivo_optimizado(file_qaqc, anio_filtro=anio_seleccionado)
            if df is not None:
                st.write("📐 Calculando desviaciones...")
                st.session_state.df_qaqc = procesar_qaqc(df, anio_seleccionado)
                status.update(label=f"✅ QAQC cargado: {len(st.session_state.df_qaqc):,} pozos", state="complete")

    st.subheader("Archivos de Planes")

    file_semanal = st.file_uploader("📅 Plan Semanal", type=['xlsx', 'xls'], key='semanal')
    if file_semanal:
        with st.status("📅 Cargando plan semanal...", expanded=True) as status:
            st.write("🗓️ Leyendo metas diarias...")
            df = pd.read_excel(file_semanal, header=0)
            st.session_state.plan_semanal = procesar_plan_semanal(df)
            status.update(label=f"✅ Plan Semanal: {len(st.session_state.plan_semanal)} días", state="complete")

    file_mensual = st.file_uploader("📆 Plan Mensual", type=['xlsx', 'xls'], key='mensual')
    if file_mensual:
        with st.status("📆 Cargando plan mensual...", expanded=True) as status:
            st.write("📊 Procesando metas mensuales...")
            df = pd.read_excel(file_mensual, header=0)
            st.session_state.plan_mensual = procesar_plan_mensual(df)
            status.update(label=f"✅ Plan Mensual: {len(st.session_state.plan_mensual)} equipos", state="complete")

# Contenido principal
if st.session_state.df_uebd is not None and st.session_state.df_qaqc is not None:

    # Inicializar filtros en session_state si no existen
    if 'filtro_equipo' not in st.session_state:
        st.session_state.filtro_equipo = 'TODOS'
    if 'fecha_ini' not in st.session_state:
        fechas_disp = sorted(st.session_state.df_qaqc['fecha'].unique())
        st.session_state.fecha_ini = min(fechas_disp) if fechas_disp else None
        st.session_state.fecha_fin = max(fechas_disp) if fechas_disp else None

    # Filtros globales en sidebar para no recargar al cambiar
    with st.sidebar:
        st.divider()
        st.subheader("🔍 Filtros de Análisis")

        equipos_disponibles = ['TODOS'] + sorted(st.session_state.df_uebd['rig'].unique().tolist())
        equipo_filtro = st.selectbox("Equipo", equipos_disponibles,
                                     index=equipos_disponibles.index(st.session_state.filtro_equipo) if st.session_state.filtro_equipo in equipos_disponibles else 0,
                                     key='sel_equipo')

        fechas_disponibles = sorted(st.session_state.df_qaqc['fecha'].unique())
        if len(fechas_disponibles) > 0:
            fecha_ini = st.date_input("Fecha inicio", value=st.session_state.fecha_ini, key='sel_fecha_ini')
            fecha_fin = st.date_input("Fecha fin", value=st.session_state.fecha_fin, key='sel_fecha_fin')

            # Botón para aplicar filtros
            if st.button("🔄 Aplicar Filtros", type="primary", use_container_width=True):
                st.session_state.filtro_equipo = equipo_filtro
                st.session_state.fecha_ini = fecha_ini
                st.session_state.fecha_fin = fecha_fin
                st.rerun()
        else:
            fecha_ini = None
            fecha_fin = None

    # Usar filtros guardados en session_state
    equipo_filtro = st.session_state.filtro_equipo
    fecha_ini = st.session_state.fecha_ini
    fecha_fin = st.session_state.fecha_fin

    # PRE-FILTRAR DATOS UNA SOLA VEZ (optimización)
    @st.cache_data
    def filtrar_datos(_df_uebd, _df_qaqc, equipo, f_ini, f_fin):
        df_u = _df_uebd
        df_q = _df_qaqc

        if equipo and equipo != 'TODOS':
            df_u = df_u[df_u['rig'] == equipo]
            df_q = df_q[df_q['rig'] == equipo]

        if f_ini and f_fin:
            df_u = df_u[(df_u['fecha'] >= f_ini) & (df_u['fecha'] <= f_fin)]
            df_q = df_q[(df_q['fecha'] >= f_ini) & (df_q['fecha'] <= f_fin)]

        return df_u, df_q

    # Filtrar datos una sola vez
    df_u_filtrado, df_q_filtrado = filtrar_datos(
        st.session_state.df_uebd,
        st.session_state.df_qaqc,
        equipo_filtro,
        fecha_ini,
        fecha_fin
    )

    # Mostrar filtros activos
    st.info(f"📊 **Filtros activos:** Equipo: {equipo_filtro} | Período: {fecha_ini} a {fecha_fin} | Registros UEBD: {len(df_u_filtrado):,} | Registros QAQC: {len(df_q_filtrado):,}")

    # Tabs principales
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📅 Metros Diarios vs Mensual",
        "📊 Real vs Semanal vs Mensual",
        "📈 Análisis Anual por Mes",
        "🔧 Detalle de Códigos",
        "📥 Reporte Turnos (Excel)"
    ])

    # TAB 1: METROS DIARIOS VS MENSUAL + DESVIACIONES QAQC
    with tab1:
        st.subheader("📅 Metros Diarios y Comparación con Plan Mensual")

        # Usar datos ya filtrados
        df_q = df_q_filtrado
        df_u = df_u_filtrado

        # PANEL DE KPIs PRINCIPALES
        metricas_periodo = calcular_metricas_uebd(df_u, df_q)

        st.markdown("### 📊 KPIs del Período Seleccionado")
        kpi1, kpi2, kpi3, kpi4, kpi5, kpi6 = st.columns(6)

        with kpi1:
            st.metric("Disponibilidad", f"{metricas_periodo['disponibilidad']:.1f}%")
        with kpi2:
            st.metric("UEBD", f"{metricas_periodo['uebd']:.1f}%")
        with kpi3:
            st.metric("Rend. ASARCO", f"{metricas_periodo['rend_asarco']:.1f} m/h")
        with kpi4:
            st.metric("Rend. Efectivo", f"{metricas_periodo['rend_efectivo']:.1f} m/h")
        with kpi5:
            st.metric("Metros Totales", f"{metricas_periodo['metros']:,.0f} m")
        with kpi6:
            st.metric("Pozos", f"{metricas_periodo['pozos']:,}")

        # Desglose de horas
        st.markdown("### ⏱️ Desglose de Tiempos")
        h = metricas_periodo['horas']
        h_cols = st.columns(7)
        h_labels = ['Efectivo', 'Perforando', 'Demora Prog', 'Demora No Prog', 'Reserva', 'Mant. Prog', 'Mant. No Prog']
        h_values = [h['ef_total'], h['ef_perf'], h['dem_prog'], h['dem_no_prog'], h['reserva'], h['mant_prog'], h['mant_no_prog']]
        h_total = metricas_periodo['horas_totales']

        for i, (col, label, val) in enumerate(zip(h_cols, h_labels, h_values)):
            pct = (val / h_total * 100) if h_total > 0 else 0
            col.metric(label, f"{val:,.1f} h", f"{pct:.1f}%")

        st.divider()

        # Métricas diarias
        metros_por_dia = df_q.groupby('fecha')['metros'].sum().reset_index()
        metros_por_dia.columns = ['Fecha', 'Metros']

        # Gráfico de metros diarios
        if len(metros_por_dia) > 0:
            col1, col2 = st.columns([2, 1])

            with col1:
                fig = px.bar(metros_por_dia, x='Fecha', y='Metros',
                            title='Metros Perforados por Día',
                            color_discrete_sequence=['#00B5B8'])

                # Agregar línea de plan mensual si existe
                if st.session_state.plan_mensual:
                    # Calcular promedio diario del plan mensual
                    plan_data = st.session_state.plan_mensual
                    if equipo_filtro != 'TODOS' and equipo_filtro in plan_data:
                        # Promedio para el equipo seleccionado
                        meses_filtro = df_q['mes'].unique()
                        metros_plan_total = sum(plan_data[equipo_filtro][m]['metros'] for m in meses_filtro if m in plan_data[equipo_filtro])
                        dias_total = len(metros_por_dia)
                        if dias_total > 0:
                            promedio_plan = metros_plan_total / dias_total
                            fig.add_hline(y=promedio_plan, line_dash="dash",
                                         line_color="#EA6B2B", annotation_text=f"Plan Mensual: {promedio_plan:.1f} m/día")
                    else:
                        # Promedio de todos los equipos
                        meses_filtro = df_q['mes'].unique()
                        metros_plan_total = 0
                        for eq in plan_data:
                            metros_plan_total += sum(plan_data[eq][m]['metros'] for m in meses_filtro if m in plan_data[eq])
                        dias_total = len(metros_por_dia)
                        if dias_total > 0:
                            promedio_plan = metros_plan_total / dias_total
                            fig.add_hline(y=promedio_plan, line_dash="dash",
                                         line_color="#EA6B2B", annotation_text=f"Plan Mensual: {promedio_plan:.1f} m/día")

                st.plotly_chart(fig, use_container_width=True)

            with col2:
                # Resumen
                st.metric("Total Metros", f"{metros_por_dia['Metros'].sum():,.1f} m")
                st.metric("Promedio Diario", f"{metros_por_dia['Metros'].mean():,.1f} m")
                st.metric("Máximo", f"{metros_por_dia['Metros'].max():,.1f} m")
                st.metric("Mínimo", f"{metros_por_dia['Metros'].min():,.1f} m")

        st.divider()

        # ANÁLISIS DE DESVIACIONES QAQC POR MALLA
        st.subheader("🎯 Análisis de Desviaciones por Malla (QAQC)")

        if 'desv_xy' in df_q.columns and 'desv_largo' in df_q.columns:
            # Filtrar registros con desviaciones válidas
            df_desv = df_q.dropna(subset=['desv_xy', 'desv_largo'])

            if len(df_desv) > 0:
                # Agrupar por malla
                desv_por_malla = df_desv.groupby('malla').agg({
                    'desv_xy': ['mean', 'std', 'count'],
                    'desv_largo': ['mean', 'std'],
                    'metros': 'sum'
                }).round(2)
                desv_por_malla.columns = ['Desv XY (prom)', 'Desv XY (std)', 'N° Pozos',
                                          'Desv Largo (prom)', 'Desv Largo (std)', 'Metros']
                desv_por_malla = desv_por_malla.reset_index()

                col1, col2 = st.columns(2)

                with col1:
                    fig_xy = px.bar(desv_por_malla, x='malla', y='Desv XY (prom)',
                                    error_y='Desv XY (std)',
                                    title='Desviación XY Promedio por Malla',
                                    color_discrete_sequence=['#3498db'])
                    fig_xy.add_hline(y=0, line_dash="solid", line_color="gray")
                    st.plotly_chart(fig_xy, use_container_width=True)

                with col2:
                    fig_largo = px.bar(desv_por_malla, x='malla', y='Desv Largo (prom)',
                                       error_y='Desv Largo (std)',
                                       title='Desviación Largo Promedio por Malla',
                                       color_discrete_sequence=['#e74c3c'])
                    fig_largo.add_hline(y=0, line_dash="solid", line_color="gray")
                    st.plotly_chart(fig_largo, use_container_width=True)

                # Tabla resumen
                st.dataframe(desv_por_malla, use_container_width=True)

                # Resumen general
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Desv XY Promedio General", f"{df_desv['desv_xy'].mean():.2f} m")
                with col2:
                    st.metric("Desv Largo Promedio General", f"{df_desv['desv_largo'].mean():.2f} m")
                with col3:
                    st.metric("Total Pozos Analizados", f"{len(df_desv):,}")
                with col4:
                    st.metric("N° de Mallas", f"{df_desv['malla'].nunique()}")
            else:
                st.info("No hay datos de desviación disponibles en el período seleccionado")
        else:
            st.warning("""
            ⚠️ **Columnas de desviación no encontradas en el archivo QAQC.**

            Para habilitar el análisis de desviaciones, asegúrese de que el archivo QAQC contenga columnas como:
            - Desviación XY / Deviation XY / Desv_Horizontal
            - Desviación Largo / Deviation Length / Desv_Profundidad
            - Malla / Mesh / Pattern
            """)

        # Tabla detallada de metros diarios
        st.subheader("📋 Detalle de Metros Diarios")

        # Agrupar por fecha y equipo
        detalle_diario = df_q.groupby(['fecha', 'rig']).agg({
            'metros': 'sum',
            'hole': 'nunique'
        }).reset_index()
        detalle_diario.columns = ['Fecha', 'Equipo', 'Metros', 'Pozos']

        # Pivot para ver equipos en columnas
        pivot_metros = detalle_diario.pivot(index='Fecha', columns='Equipo', values='Metros').fillna(0)
        pivot_metros['Total'] = pivot_metros.sum(axis=1)

        st.dataframe(pivot_metros.style.format("{:.1f}"), use_container_width=True)

    # TAB 2: REAL VS SEMANAL VS MENSUAL
    with tab2:
        st.subheader("📊 Comparación: Real vs Plan Semanal vs Plan Mensual")

        if st.session_state.plan_semanal is None and st.session_state.plan_mensual is None:
            st.warning("⚠️ Cargue al menos un archivo de plan (semanal o mensual) para ver esta comparación")
        else:
            # Selector de métrica
            metrica_sel = st.selectbox("Seleccione la métrica a comparar",
                                       ['metros', 'disponibilidad', 'utilizacion', 'rendimiento', 'horas_efectivas'],
                                       format_func=lambda x: {
                                           'metros': 'Metros',
                                           'disponibilidad': 'Disponibilidad %',
                                           'utilizacion': 'Utilización %',
                                           'rendimiento': 'Rendimiento m/h',
                                           'horas_efectivas': 'Horas Efectivas'
                                       }.get(x, x))

            # Usar datos ya filtrados
            df_q = df_q_filtrado
            df_u = df_u_filtrado

            fechas_rango = sorted(df_q['fecha'].unique())

            # Calcular valores reales, semanales y mensuales por fecha
            comparacion = []

            for fecha in fechas_rango:
                df_q_dia = df_q[df_q['fecha'] == fecha]
                df_u_dia = df_u[df_u['fecha'] == fecha]

                # Calcular métricas reales del día
                metricas_real = calcular_metricas_uebd(df_u_dia, df_q_dia)

                # Valor real
                if metrica_sel == 'metros':
                    val_real = df_q_dia['metros'].sum()
                elif metrica_sel == 'disponibilidad':
                    val_real = metricas_real['disponibilidad']
                elif metrica_sel == 'utilizacion':
                    val_real = metricas_real['uebd']
                elif metrica_sel == 'rendimiento':
                    val_real = metricas_real['rend_asarco']
                elif metrica_sel == 'horas_efectivas':
                    val_real = metricas_real['horas']['ef_total']
                else:
                    val_real = 0

                # Valor plan semanal
                val_semanal = 0
                if st.session_state.plan_semanal and fecha in st.session_state.plan_semanal:
                    if equipo_filtro != 'TODOS' and equipo_filtro in st.session_state.plan_semanal[fecha]['por_rig']:
                        val_semanal = st.session_state.plan_semanal[fecha]['por_rig'][equipo_filtro].get(metrica_sel, 0)
                    else:
                        val_semanal = st.session_state.plan_semanal[fecha]['total'].get(metrica_sel, 0)

                # Valor plan mensual (convertido a diario)
                val_mensual = 0
                if st.session_state.plan_mensual:
                    mes = fecha.month
                    import calendar
                    dias_mes = calendar.monthrange(fecha.year, mes)[1]

                    if equipo_filtro != 'TODOS' and equipo_filtro in st.session_state.plan_mensual:
                        val_mes = st.session_state.plan_mensual[equipo_filtro].get(mes, {}).get(metrica_sel, 0)
                        if metrica_sel == 'metros' or metrica_sel == 'horas_efectivas':
                            val_mensual = val_mes / dias_mes  # Dividir entre días del mes
                        else:
                            val_mensual = val_mes  # Porcentajes y rendimientos no se dividen
                    else:
                        # Sumar todos los equipos
                        for eq_data in st.session_state.plan_mensual.values():
                            val_mes = eq_data.get(mes, {}).get(metrica_sel, 0)
                            if metrica_sel == 'metros' or metrica_sel == 'horas_efectivas':
                                val_mensual += val_mes / dias_mes
                            else:
                                val_mensual += val_mes
                        if metrica_sel not in ['metros', 'horas_efectivas']:
                            val_mensual /= len(st.session_state.plan_mensual)  # Promedio para porcentajes

                comparacion.append({
                    'Fecha': fecha,
                    'Real': val_real,
                    'Plan Semanal': val_semanal,
                    'Plan Mensual': val_mensual,
                    'Diff Real-Semanal': val_real - val_semanal,
                    'Diff Real-Mensual': val_real - val_mensual,
                    'Diff Semanal-Mensual': val_semanal - val_mensual
                })

            df_comp = pd.DataFrame(comparacion)

            if len(df_comp) > 0:
                # Gráfico de comparación
                fig = go.Figure()

                fig.add_trace(go.Scatter(x=df_comp['Fecha'], y=df_comp['Real'],
                                        mode='lines+markers', name='Real',
                                        line=dict(color='#455A64', width=2)))

                if st.session_state.plan_semanal:
                    fig.add_trace(go.Scatter(x=df_comp['Fecha'], y=df_comp['Plan Semanal'],
                                            mode='lines+markers', name='Plan Semanal',
                                            line=dict(color='#0277BD', width=2, dash='dash')))

                if st.session_state.plan_mensual:
                    fig.add_trace(go.Scatter(x=df_comp['Fecha'], y=df_comp['Plan Mensual'],
                                            mode='lines+markers', name='Plan Mensual',
                                            line=dict(color='#7B1FA2', width=2, dash='dot')))

                fig.update_layout(title=f'Comparación: {metrica_sel.replace("_", " ").title()}',
                                 xaxis_title='Fecha', yaxis_title='Valor',
                                 hovermode='x unified')

                st.plotly_chart(fig, use_container_width=True)

                # Métricas resumen
                col1, col2, col3 = st.columns(3)

                with col1:
                    total_real = df_comp['Real'].sum() if metrica_sel in ['metros', 'horas_efectivas'] else df_comp['Real'].mean()
                    st.metric("Real (Total/Prom)", f"{total_real:,.1f}")

                with col2:
                    total_semanal = df_comp['Plan Semanal'].sum() if metrica_sel in ['metros', 'horas_efectivas'] else df_comp['Plan Semanal'].mean()
                    diff_rs = total_real - total_semanal
                    st.metric("Plan Semanal", f"{total_semanal:,.1f}",
                             delta=f"{diff_rs:+,.1f}" if total_semanal > 0 else None)

                with col3:
                    total_mensual = df_comp['Plan Mensual'].sum() if metrica_sel in ['metros', 'horas_efectivas'] else df_comp['Plan Mensual'].mean()
                    diff_rm = total_real - total_mensual
                    st.metric("Plan Mensual", f"{total_mensual:,.1f}",
                             delta=f"{diff_rm:+,.1f}" if total_mensual > 0 else None)

                # Tabla de comparación
                st.subheader("📋 Tabla Detallada")

                # Formatear tabla
                df_display = df_comp.copy()
                df_display['Fecha'] = df_display['Fecha'].astype(str)

                st.dataframe(df_display.style.format({
                    'Real': '{:.1f}',
                    'Plan Semanal': '{:.1f}',
                    'Plan Mensual': '{:.1f}',
                    'Diff Real-Semanal': '{:+.1f}',
                    'Diff Real-Mensual': '{:+.1f}',
                    'Diff Semanal-Mensual': '{:+.1f}'
                }).applymap(lambda x: 'color: green' if isinstance(x, (int, float)) and x > 0 else 'color: red' if isinstance(x, (int, float)) and x < 0 else '',
                           subset=['Diff Real-Semanal', 'Diff Real-Mensual', 'Diff Semanal-Mensual']),
                use_container_width=True)

    # TAB 3: ANÁLISIS ANUAL POR MES
    with tab3:
        st.subheader("📈 Análisis Anual por Mes")

        # Para análisis anual, usar todos los datos del año (solo filtro de equipo)
        df_q = st.session_state.df_qaqc
        df_u = st.session_state.df_uebd

        if equipo_filtro != 'TODOS':
            df_q = df_q[df_q['rig'] == equipo_filtro]
            df_u = df_u[df_u['rig'] == equipo_filtro]

        # Calcular métricas por mes
        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        metricas_por_mes = []

        for mes in range(1, 13):
            df_q_mes = df_q[df_q['mes'] == mes]
            df_u_mes = df_u[df_u['mes'] == mes]

            if len(df_q_mes) > 0 or len(df_u_mes) > 0:
                metricas = calcular_metricas_uebd(df_u_mes, df_q_mes)

                # Agregar valores del plan mensual si existe
                plan_metros = 0
                plan_disp = 0
                plan_util = 0
                plan_rend = 0

                if st.session_state.plan_mensual:
                    if equipo_filtro != 'TODOS' and equipo_filtro in st.session_state.plan_mensual:
                        plan_data = st.session_state.plan_mensual[equipo_filtro].get(mes, {})
                        plan_metros = plan_data.get('metros', 0)
                        plan_disp = plan_data.get('disponibilidad', 0)
                        plan_util = plan_data.get('utilizacion', 0)
                        plan_rend = plan_data.get('rendimiento', 0)
                    else:
                        # Sumar todos los equipos
                        count = 0
                        for eq_data in st.session_state.plan_mensual.values():
                            if mes in eq_data:
                                plan_metros += eq_data[mes].get('metros', 0)
                                plan_disp += eq_data[mes].get('disponibilidad', 0)
                                plan_util += eq_data[mes].get('utilizacion', 0)
                                plan_rend += eq_data[mes].get('rendimiento', 0)
                                count += 1
                        if count > 0:
                            plan_disp /= count
                            plan_util /= count
                            plan_rend /= count

                metricas_por_mes.append({
                    'Mes': meses_nombres[mes],
                    'N° Mes': mes,
                    'Disponibilidad %': metricas['disponibilidad'],
                    'Plan Disp %': plan_disp,
                    'UEBD %': metricas['uebd'],
                    'Plan UEBD %': plan_util,
                    'Rend. ASARCO': metricas['rend_asarco'],
                    'Plan Rend.': plan_rend,
                    'Metros': metricas['metros'],
                    'Plan Metros': plan_metros,
                    'Hrs Efectivas': metricas['horas']['ef_total'],
                    'Hrs Perforando': metricas['horas']['ef_perf'],
                    'Pozos': metricas['pozos'],
                    'Hrs Demora Prog': metricas['horas']['dem_prog'],
                    'Hrs Demora No Prog': metricas['horas']['dem_no_prog'],
                    'Hrs Mantencion': metricas['horas']['mant_prog'] + metricas['horas']['mant_no_prog']
                })

        df_meses = pd.DataFrame(metricas_por_mes)

        if len(df_meses) > 0:
            # Gráficos por mes
            col1, col2 = st.columns(2)

            with col1:
                fig1 = go.Figure()
                fig1.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Metros'],
                                     name='Real', marker_color='#00B5B8'))
                if st.session_state.plan_mensual:
                    fig1.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Plan Metros'],
                                         name='Plan', marker_color='#EA6B2B', opacity=0.7))
                fig1.update_layout(title='Metros por Mes', barmode='group')
                st.plotly_chart(fig1, use_container_width=True)

            with col2:
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=df_meses['Mes'], y=df_meses['Disponibilidad %'],
                                         mode='lines+markers', name='Disponibilidad Real',
                                         line=dict(color='#00B5B8')))
                fig2.add_trace(go.Scatter(x=df_meses['Mes'], y=df_meses['UEBD %'],
                                         mode='lines+markers', name='UEBD Real',
                                         line=dict(color='#3498db')))
                if st.session_state.plan_mensual:
                    fig2.add_trace(go.Scatter(x=df_meses['Mes'], y=df_meses['Plan Disp %'],
                                             mode='lines', name='Disponibilidad Plan',
                                             line=dict(color='#00B5B8', dash='dash')))
                    fig2.add_trace(go.Scatter(x=df_meses['Mes'], y=df_meses['Plan UEBD %'],
                                             mode='lines', name='UEBD Plan',
                                             line=dict(color='#3498db', dash='dash')))
                fig2.update_layout(title='Disponibilidad y UEBD por Mes', yaxis_title='%')
                st.plotly_chart(fig2, use_container_width=True)

            # Gráfico de rendimientos
            col1, col2 = st.columns(2)

            with col1:
                fig3 = go.Figure()
                fig3.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Rend. ASARCO'],
                                     name='Real', marker_color='#00B5B8'))
                if st.session_state.plan_mensual:
                    fig3.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Plan Rend.'],
                                         name='Plan', marker_color='#EA6B2B', opacity=0.7))
                fig3.update_layout(title='Rendimiento ASARCO por Mes', barmode='group',
                                  yaxis_title='m/h')
                st.plotly_chart(fig3, use_container_width=True)

            with col2:
                # Gráfico de demoras
                fig4 = go.Figure()
                fig4.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Hrs Demora Prog'],
                                     name='Demora Programada', marker_color='#f39c12'))
                fig4.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Hrs Demora No Prog'],
                                     name='Demora No Programada', marker_color='#e74c3c'))
                fig4.add_trace(go.Bar(x=df_meses['Mes'], y=df_meses['Hrs Mantencion'],
                                     name='Mantención', marker_color='#3498db'))
                fig4.update_layout(title='Desglose de Tiempos No Productivos por Mes',
                                  barmode='stack', yaxis_title='Horas')
                st.plotly_chart(fig4, use_container_width=True)

            # Tabla resumen anual
            st.subheader("📋 Tabla Resumen Anual")

            # Formatear columnas para mostrar
            cols_mostrar = ['Mes', 'Disponibilidad %', 'UEBD %', 'Rend. ASARCO',
                           'Metros', 'Hrs Efectivas', 'Pozos']
            if st.session_state.plan_mensual:
                cols_mostrar = ['Mes', 'Disponibilidad %', 'Plan Disp %', 'UEBD %', 'Plan UEBD %',
                               'Rend. ASARCO', 'Plan Rend.', 'Metros', 'Plan Metros',
                               'Hrs Efectivas', 'Pozos']

            df_display = df_meses[cols_mostrar].copy()

            # Agregar fila de totales/promedios
            totales = {'Mes': 'TOTAL/PROM'}
            for col in cols_mostrar[1:]:
                if 'Metros' in col or 'Hrs' in col or 'Pozos' in col:
                    totales[col] = df_display[col].sum()
                else:
                    totales[col] = df_display[col].mean()

            df_display = pd.concat([df_display, pd.DataFrame([totales])], ignore_index=True)

            st.dataframe(df_display.style.format({
                'Disponibilidad %': '{:.1f}%',
                'Plan Disp %': '{:.1f}%',
                'UEBD %': '{:.1f}%',
                'Plan UEBD %': '{:.1f}%',
                'Rend. ASARCO': '{:.1f}',
                'Plan Rend.': '{:.1f}',
                'Metros': '{:,.0f}',
                'Plan Metros': '{:,.0f}',
                'Hrs Efectivas': '{:,.1f}',
                'Pozos': '{:,.0f}'
            }), use_container_width=True)

            # Análisis por equipo si no hay filtro
            if equipo_filtro == 'TODOS':
                st.subheader("📊 Análisis por Equipo")

                equipos = sorted(st.session_state.df_qaqc['rig'].unique())
                metricas_equipo = []

                for eq in equipos:
                    df_q_eq = st.session_state.df_qaqc[st.session_state.df_qaqc['rig'] == eq]
                    df_u_eq = st.session_state.df_uebd[st.session_state.df_uebd['rig'] == eq]

                    metricas = calcular_metricas_uebd(df_u_eq, df_q_eq)

                    metricas_equipo.append({
                        'Equipo': eq,
                        'Disponibilidad %': metricas['disponibilidad'],
                        'UEBD %': metricas['uebd'],
                        'Rend. ASARCO': metricas['rend_asarco'],
                        'Metros': metricas['metros'],
                        'Pozos': metricas['pozos'],
                        'Hrs Efectivas': metricas['horas']['ef_total']
                    })

                df_equipos = pd.DataFrame(metricas_equipo)

                col1, col2 = st.columns(2)

                with col1:
                    fig_eq1 = px.bar(df_equipos, x='Equipo', y='Metros',
                                    title='Metros por Equipo',
                                    color_discrete_sequence=['#00B5B8'])
                    st.plotly_chart(fig_eq1, use_container_width=True)

                with col2:
                    fig_eq2 = px.bar(df_equipos, x='Equipo', y='Rend. ASARCO',
                                    title='Rendimiento ASARCO por Equipo',
                                    color_discrete_sequence=['#EA6B2B'])
                    st.plotly_chart(fig_eq2, use_container_width=True)

                st.dataframe(df_equipos.style.format({
                    'Disponibilidad %': '{:.1f}%',
                    'UEBD %': '{:.1f}%',
                    'Rend. ASARCO': '{:.1f}',
                    'Metros': '{:,.0f}',
                    'Pozos': '{:,.0f}',
                    'Hrs Efectivas': '{:,.1f}'
                }), use_container_width=True)

    # TAB 4: DETALLE DE CÓDIGOS
    with tab4:
        st.subheader("🔧 Detalle de Códigos y Tiempos")

        # Usar datos ya filtrados (hacer copia para modificar)
        df_u = df_u_filtrado.copy()

        if len(df_u) > 0:
            # Análisis por estado (ShortCode)
            st.subheader("📊 Distribución por Estado (ShortCode)")

            df_u['horas'] = df_u['duracion'] / 3600

            por_estado = df_u.groupby('estado')['horas'].sum().reset_index()
            por_estado.columns = ['Estado', 'Horas']
            por_estado = por_estado.sort_values('Horas', ascending=False)

            col1, col2 = st.columns([1, 1])

            with col1:
                fig_estado = px.pie(por_estado, values='Horas', names='Estado',
                                   title='Distribución de Horas por Estado',
                                   color_discrete_sequence=px.colors.qualitative.Set2)
                st.plotly_chart(fig_estado, use_container_width=True)

            with col2:
                fig_estado_bar = px.bar(por_estado, x='Estado', y='Horas',
                                       title='Horas por Estado',
                                       color_discrete_sequence=['#00B5B8'])
                st.plotly_chart(fig_estado_bar, use_container_width=True)

            st.dataframe(por_estado.style.format({'Horas': '{:,.1f}'}), use_container_width=True)

            st.divider()

            # Análisis por código (CodeName)
            st.subheader("📋 Detalle por Código (CodeName)")

            por_codigo = df_u.groupby(['estado', 'codigo'])['horas'].sum().reset_index()
            por_codigo.columns = ['Estado', 'Código', 'Horas']
            por_codigo = por_codigo.sort_values(['Estado', 'Horas'], ascending=[True, False])

            # Filtro por estado
            estados_unicos = ['Todos'] + sorted(df_u['estado'].unique().tolist())
            estado_filtro = st.selectbox("Filtrar por Estado", estados_unicos)

            if estado_filtro != 'Todos':
                por_codigo = por_codigo[por_codigo['Estado'] == estado_filtro]

            # Mostrar top códigos
            top_n = st.slider("Mostrar top N códigos", 5, 50, 20)

            por_codigo_top = por_codigo.nlargest(top_n, 'Horas')

            fig_codigos = px.bar(por_codigo_top, x='Horas', y='Código', orientation='h',
                                color='Estado', title=f'Top {top_n} Códigos por Horas',
                                height=max(400, top_n * 25))
            fig_codigos.update_layout(yaxis={'categoryorder': 'total ascending'})
            st.plotly_chart(fig_codigos, use_container_width=True)

            st.dataframe(por_codigo_top.style.format({'Horas': '{:,.2f}'}), use_container_width=True)

            st.divider()

            # Análisis de planificación
            st.subheader("📅 Análisis Programado vs No Programado")

            df_u['tipo_planificacion'] = df_u['planificado'].apply(
                lambda x: 'Programado' if x == 'Programada' else 'No Programado' if pd.notna(x) and x else 'Sin Clasificar'
            )

            por_planificacion = df_u.groupby(['estado', 'tipo_planificacion'])['horas'].sum().reset_index()
            por_planificacion.columns = ['Estado', 'Tipo', 'Horas']

            fig_plan = px.sunburst(por_planificacion, path=['Estado', 'Tipo'], values='Horas',
                                   title='Distribución de Horas: Estado → Programado/No Programado',
                                   color_discrete_sequence=px.colors.qualitative.Pastel)
            st.plotly_chart(fig_plan, use_container_width=True)

            # Tabla resumen por estado y planificación
            pivot_plan = por_planificacion.pivot(index='Estado', columns='Tipo', values='Horas').fillna(0)
            pivot_plan['Total'] = pivot_plan.sum(axis=1)

            st.dataframe(pivot_plan.style.format('{:,.1f}'), use_container_width=True)

            st.divider()

            # Evolución temporal de códigos
            st.subheader("📈 Evolución Temporal por Código")

            # Seleccionar códigos a analizar
            codigos_principales = por_codigo.nlargest(10, 'Horas')['Código'].tolist()
            codigos_seleccionados = st.multiselect("Seleccione códigos a analizar",
                                                   codigos_principales,
                                                   default=codigos_principales[:3])

            if codigos_seleccionados:
                df_evol = df_u[df_u['codigo'].isin(codigos_seleccionados)]
                evol_diaria = df_evol.groupby(['fecha', 'codigo'])['horas'].sum().reset_index()

                fig_evol = px.line(evol_diaria, x='fecha', y='horas', color='codigo',
                                  title='Evolución Diaria de Horas por Código Seleccionado')
                st.plotly_chart(fig_evol, use_container_width=True)

    # TAB 5: REPORTE TURNOS (EXCEL)
    with tab5:
        st.subheader("📥 Reporte de Turnos (TA/TB) - Exportable a Excel")
        st.caption("Este reporte mantiene los metros como números para que puedas sumar/filtrar en Excel.")

        # Preparar QAQC con turno y día operacional
        df_q_rep = agregar_dia_operacional_y_turno(df_q_filtrado, "timestamp", out_date_col="dia_operacional")

        if len(df_q_rep) == 0:
            st.info("No hay datos QAQC en el rango seleccionado para generar el reporte.")
        else:
            # Selector de día operacional
            dias = sorted(pd.Series(df_q_rep["dia_operacional"]).dropna().unique().tolist())
            dia_sel = st.selectbox("Seleccione día operacional", dias, index=len(dias) - 1 if len(dias) > 0 else 0)

            df_q_dia = df_q_rep[df_q_rep["dia_operacional"] == dia_sel].copy()

            # Frente (si existe en malla/pattern)
            if "malla" in df_q_dia.columns:
                df_q_dia["frente"] = df_q_dia["malla"].astype(str).str.extract(r"(F\\d{2})", expand=False).fillna("")
            else:
                df_q_dia["frente"] = ""

            # Metros por rig y turno
            piv = (
                df_q_dia.groupby(["frente", "rig", "turno"], dropna=False)["metros"]
                .sum()
                .unstack("turno", fill_value=0)
                .reset_index()
            )
            if "A" not in piv.columns:
                piv["A"] = 0.0
            if "B" not in piv.columns:
                piv["B"] = 0.0

            piv["TURNO A"] = pd.to_numeric(piv["A"], errors="coerce").fillna(0.0)
            piv["TURNO B"] = pd.to_numeric(piv["B"], errors="coerce").fillna(0.0)
            piv = piv.drop(columns=[c for c in ["A", "B"] if c in piv.columns])
            piv["TOTAL"] = piv["TURNO A"] + piv["TURNO B"]

            # PLAN (si hay plan semanal cargado)
            plan_dict = st.session_state.plan_semanal or {}

            def _plan_metros(fecha_op, rig):
                try:
                    f = fecha_op
                    if isinstance(fecha_op, str):
                        f = pd.to_datetime(fecha_op).date()
                    d = plan_dict.get(f, {}).get("por_rig", {}).get(rig, {})
                    return float(d.get("metros", 0) or 0)
                except:
                    return 0.0

            piv["PLAN"] = piv["rig"].apply(lambda r: _plan_metros(dia_sel, r))

            # Cumplimientos (%)
            def _pct(num, den):
                return (num / den * 100.0) if den and den > 0 else 0.0

            piv["CUMPLIMIENTO TA"] = piv.apply(lambda row: _pct(row["TURNO A"], row["PLAN"]), axis=1)
            piv["CUMPLIMIENTO TB"] = piv.apply(lambda row: _pct(row["TURNO B"], row["PLAN"]), axis=1)
            piv["CUMPLIMIENTO DIARIO"] = piv.apply(lambda row: _pct(row["TOTAL"], row["PLAN"]), axis=1)

            # ESTADO PERFORADORA (desde UEBD, por turno)
            df_u_rep = agregar_dia_operacional_y_turno(df_u_filtrado, "timestamp", out_date_col="dia_operacional")
            estado_piv = None
            if len(df_u_rep) > 0 and {"rig", "estado", "duracion", "turno", "dia_operacional"}.issubset(df_u_rep.columns):
                df_u_dia = df_u_rep[df_u_rep["dia_operacional"] == dia_sel].copy()
                if len(df_u_dia) > 0:
                    df_u_dia["horas"] = pd.to_numeric(df_u_dia["duracion"], errors="coerce").fillna(0.0) / 3600.0
                    tmp = (
                        df_u_dia.groupby(["rig", "turno", "estado"], dropna=False)["horas"]
                        .sum()
                        .reset_index()
                    )
                    if len(tmp) > 0:
                        # Estado dominante por rig+turno
                        tmp = tmp.sort_values(["rig", "turno", "horas"], ascending=[True, True, False])
                        dom = tmp.drop_duplicates(subset=["rig", "turno"], keep="first")
                        estado_piv = dom.pivot(index="rig", columns="turno", values="estado").reset_index()
                        if "A" not in estado_piv.columns:
                            estado_piv["A"] = ""
                        if "B" not in estado_piv.columns:
                            estado_piv["B"] = ""
                        estado_piv = estado_piv.rename(columns={"A": "ESTADO PERFORADORA TA", "B": "ESTADO PERFORADORA TB"})

            if estado_piv is not None:
                piv = piv.merge(estado_piv, on="rig", how="left")
            else:
                piv["ESTADO PERFORADORA TA"] = ""
                piv["ESTADO PERFORADORA TB"] = ""

            # Orden columnas
            cols_final = [
                "frente", "rig",
                "TURNO A", "TURNO B", "TOTAL", "PLAN",
                "CUMPLIMIENTO TA", "CUMPLIMIENTO TB", "CUMPLIMIENTO DIARIO",
                "ESTADO PERFORADORA TA", "ESTADO PERFORADORA TB",
            ]
            for c in cols_final:
                if c not in piv.columns:
                    piv[c] = ""
            piv = piv[cols_final].copy()

            # Totales por frente + total global
            def _total_row(df_sub, frente_label):
                sA = float(df_sub["TURNO A"].sum())
                sB = float(df_sub["TURNO B"].sum())
                sT = float(df_sub["TOTAL"].sum())
                sP = float(df_sub["PLAN"].sum())
                return {
                    "frente": frente_label,
                    "rig": "TOTAL",
                    "TURNO A": sA,
                    "TURNO B": sB,
                    "TOTAL": sT,
                    "PLAN": sP,
                    "CUMPLIMIENTO TA": _pct(sA, sP),
                    "CUMPLIMIENTO TB": _pct(sB, sP),
                    "CUMPLIMIENTO DIARIO": _pct(sT, sP),
                    "ESTADO PERFORADORA TA": "",
                    "ESTADO PERFORADORA TB": "",
                }

            frames = []
            frentes = piv["frente"].fillna("").unique().tolist()
            frentes_orden = sorted([f for f in frentes if f != ""]) + ([""] if "" in frentes else [])
            for f in frentes_orden:
                sub = piv[piv["frente"] == f].copy()
                sub = sub.sort_values(["rig"])
                frames.append(sub)
                frames.append(pd.DataFrame([_total_row(sub, f if f else "SIN_FRENTE")]))

            df_reporte = pd.concat(frames, ignore_index=True) if frames else piv.copy()

            # Total global
            total_global = _total_row(piv, "TOTAL METROS")
            total_global["rig"] = "TOTAL METROS"
            df_reporte = pd.concat([df_reporte, pd.DataFrame([total_global])], ignore_index=True)

            # Mostrar en pantalla (formato)
            df_show = df_reporte.copy()
            st.dataframe(
                df_show.style.format({
                    "TURNO A": "{:,.0f}",
                    "TURNO B": "{:,.0f}",
                    "TOTAL": "{:,.0f}",
                    "PLAN": "{:,.0f}",
                    "CUMPLIMIENTO TA": "{:.1f}%",
                    "CUMPLIMIENTO TB": "{:.1f}%",
                    "CUMPLIMIENTO DIARIO": "{:.1f}%",
                }),
                use_container_width=True
            )

            # Export a Excel (mantener numéricos)
            export_df = df_reporte.copy()
            # Asegurar tipos numéricos
            for c in ["TURNO A", "TURNO B", "TOTAL", "PLAN", "CUMPLIMIENTO TA", "CUMPLIMIENTO TB", "CUMPLIMIENTO DIARIO"]:
                export_df[c] = pd.to_numeric(export_df[c], errors="coerce").fillna(0.0)

            output = io.BytesIO()
            try:
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    export_df.to_excel(writer, index=False, sheet_name="Turnos")

                    # Formato básico en Excel (porcentaje)
                    ws = writer.book["Turnos"]
                    header = [cell.value for cell in ws[1]]
                    pct_cols = {"CUMPLIMIENTO TA", "CUMPLIMIENTO TB", "CUMPLIMIENTO DIARIO"}
                    for j, name in enumerate(header, start=1):
                        if name in pct_cols:
                            for i in range(2, ws.max_row + 1):
                                ws.cell(row=i, column=j).number_format = "0.0\\%"
                        if name in {"TURNO A", "TURNO B", "TOTAL", "PLAN"}:
                            for i in range(2, ws.max_row + 1):
                                ws.cell(row=i, column=j).number_format = "#,##0"
            except Exception as e:
                st.error(f"Error exportando Excel: {e}")
            else:
                st.download_button(
                    label="⬇️ Descargar Excel (Turnos)",
                    data=output.getvalue(),
                    file_name=f"Reporte_Turnos_{dia_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

else:
    # Mensaje cuando no hay datos cargados
    st.info("""
    👋 **Bienvenido al Dashboard PyT Unificado**

    Para comenzar, cargue los archivos de datos en la barra lateral:

    1. **UEBD** - Archivo de utilización con tiempos (obligatorio)
    2. **QAQC** - Archivo de metros perforados (obligatorio)
    3. **Plan Semanal** - Plan con valores diarios (opcional)
    4. **Plan Mensual** - Plan con valores mensuales (opcional)

    Una vez cargados los archivos, podrá visualizar:
    - 📅 Metros diarios vs Plan mensual + Desviaciones QAQC
    - 📊 Comparación Real vs Semanal vs Mensual
    - 📈 Análisis anual desglosado por mes
    - 🔧 Detalle de códigos y clasificación de tiempos
    """)

    # Mostrar estructura esperada de archivos
    with st.expander("📋 Estructura esperada de archivos"):
        st.markdown("""
        ### Archivo UEBD
        Columnas principales (índice base):
        - Col 0: RigName (Equipo)
        - Col 7: WorkDayStarted (Fecha)
        - Col 14: Duration (Duración en segundos)
        - Col 15: CodeName (Código de actividad)
        - Col 16: ShortCode (Estado: Efectivo, Demora, Mantencion, Reserva)
        - Col 19: PlannedCodeName (Programada/No Programada)

        ### Archivo QAQC
        Columnas principales:
        - Col 0: RigName (Equipo)
        - Col 2: HoleName (Nombre del pozo)
        - Col 26: RealLength (Metros perforados)
        - Col 29: WorkDayStarted (Fecha)
        - Columnas opcionales para desviaciones:
          - Malla/Mesh/Pattern
          - Desviación XY
          - Desviación Largo

        ### Plan Semanal
        - Fila 0: Encabezados con fechas
        - Col 0: Equipo
        - Col 1: Índices (Disponibilidad, Utilización, Rendimiento, Metros, Horas Efectivas)
        - Col 2+: Valores por fecha

        ### Plan Mensual
        - Col 0: Equipo
        - Col 1: Índices
        - Col 2: Unidad
        - Col 3-14: Valores mensuales (Enero a Diciembre)
        """)

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #718096; font-size: 12px;'>
    Dashboard PyT Unificado v1.0 | Los Pelambres
</div>
<div style='text-align: center; margin-top: 30px; padding: 20px; background: linear-gradient(135deg, #1a365d, #2c5282); border-radius: 15px; color: white;'>
    <p style='font-size: 14px; font-weight: 600; margin-bottom: 10px;'>
        HECHO POR MARIA IGNACIA FERNANDEZ MEJOR CONOCIDA COMO PILI
    </p>
    <p style='font-size: 13px; color: #90cdf4;'>
        Los echo de menos, se cuidan por favor 💙
    </p>
</div>
""", unsafe_allow_html=True)
